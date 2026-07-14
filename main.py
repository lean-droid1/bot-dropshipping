"""
Bot de Dropshipping — Tienda Negocio / RXZ Web
Versión limpia con API WooCommerce del proveedor.
"""
import os, time, json, re, io, threading, imaplib, email
try:
    from flask import Flask, request, jsonify
    FLASK_OK = True
except ImportError:
    FLASK_OK = False
    print('⚠️ Flask no instalado — webhooks desactivados')

try:
    from curl_cffi import requests as cf_requests
    CURL_CFFI_OK = True
    print('✅ curl-cffi disponible — bypass Cloudflare activo')
except ImportError:
    cf_requests = None
    CURL_CFFI_OK = False
    print('⚠️ curl-cffi no instalado — usando requests normal')
from email.header import decode_header
from datetime import datetime, timedelta
import requests
urllib3_imported = False
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    urllib3_imported = True
except Exception:
    pass

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════
DB_FILE          = "estado_productos.json"
API_BASE         = "https://developers.tiendanegocio.com/v1"
PROV_API         = "https://rxzweb.com/wp-json/wc/store/v1/products"
USER_AGENT       = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

# Márgenes y alertas (configurables via Railway Variables)
MARGEN           = float(os.environ.get("MARGEN", "0.78"))   # /0.78 = 22% ganancia
ALERTA_STOCK       = 3       # Alerta Telegram cuando stock llega a este numero
ENVIO_GRATIS_MIN   = 100000  # Activar envio gratis en productos >= este precio

# Mesas pesadas (no envio gratis): RT-01D:3643182, RF-RT02D:3643163, LW-A1:3643153, LW-A1 Mini:3643173
# Estos productos tienen envío caro (~$200.000), se les pone cartel manual en la foto
PRODUCTOS_PESADOS_IDS = {3643182, 3643163, 3643153, 3643173}

CICLO_MINUTOS    = 15   # Cada cuántos minutos monitorea

# Categorías del proveedor a excluir siempre (no nos interesan)
CATEGORIAS_EXCLUIDAS = {'pantallas', 'modulos', 'baterias', 'bateria'}

PALABRAS_INTERES = [
    'ma ant','amaoe','2uul','goot wick','mijing','louwei','rf4','jakemy',
    'kailiwei','kslid','aifen','sugon','jcid','jc','v1','v1s','v1se',
    'v1 pro','programadora','organizador','cinta','silla','mesa','puas',
    'hilo','cepillo','flux','malla','estaño','pinza','alicate','tweezer',
    'brusela','stencil','pasta','mascara','ventosa','rodillo','soporte',
    'holder','microscopio','fuente','estacion','plancha','precalentadora',
    'autoclave','cabina','compresor','camara','detector','tester','probador'
]

# ── Variables de entorno ──────────────────────────────────────────────────────
def _e(k): return (os.environ.get(k) or "").strip() or None

TELEGRAM_TOKEN = _e("TELEGRAM_TOKEN")
CHAT_ID        = _e("CHAT_ID")
GMAIL_USER     = _e("GMAIL_USER")
GMAIL_PASS     = _e("GMAIL_PASS")
CLIENT_ID      = _e("CLIENT_ID")
CLIENT_SECRET  = _e("CLIENT_SECRET")
NOMBRE_TIENDA  = _e("NOMBRE_TIENDA") or "🧪 PRUEBA"

# ── Compras automáticas al proveedor ─────────────────────────────────────────
PROV_USER          = _e("PROV_USER")
WEBHOOK_BASE_URL   = _e("WEBHOOK_BASE_URL")
GROQ_API_KEY       = _e("GROQ_API_KEY")
PROV_PASS      = _e("PROV_PASS")
CUIT_PROVEEDOR = _e("CUIT_PROVEEDOR")
PROV_LOGIN_URL = "https://rxzweb.com/wp-login.php"
PROV_CART_URL  = "https://rxzweb.com/wp-json/wc/store/v1/cart"
PROV_CHKOUT_URL= "https://rxzweb.com/wp-json/wc/store/v1/checkout"

# ── Token API Tienda Negocio ──────────────────────────────────────────────────
_token    = None
_store_id = None

def cargar_token():
    global _token, _store_id
    t = _e("API_TOKEN"); u = _e("API_USER_ID")
    if t and u:
        _token = t; _store_id = u
        print(f"✅ Token desde Railway (store_id={u})"); return
    db = leer_db()
    if db.get("api_token") and db.get("api_user_id"):
        _token = db["api_token"]; _store_id = db["api_user_id"]
        print(f"✅ Token desde DB (store_id={_store_id})")

def guardar_token(token, store_id):
    global _token, _store_id
    _token = token; _store_id = store_id
    db = leer_db(); db["api_token"] = token; db["api_user_id"] = store_id
    escribir_db(db); print(f"💾 Token guardado (store_id={store_id})")

# ══════════════════════════════════════════════════════════════════════════════
# BASE DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
def leer_db():
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
                d.setdefault("productos_proveedor", {})
                d.setdefault("sincronizados", {})
                d.setdefault("pedidos_procesados", [])
                d.setdefault("ofertas_pendientes", {})
                d.setdefault("ordenes", {})
                d.setdefault("variantes_cache", {})
                d.setdefault("pids_refrescar", [])
                d.setdefault("ultimo_orden_id", 0)
                return d
        except Exception: pass
    return {"productos_proveedor":{}, "sincronizados":{}, "pedidos_procesados":[], "ofertas_pendientes":{}, "ordenes":{}}

def escribir_db(d):
    try:
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(d, f, ensure_ascii=False, indent=2)
    except Exception as e: print(f"❌ DB: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# TELEGRAM
# ══════════════════════════════════════════════════════════════════════════════
def tg(msg):
    if not msg or not TELEGRAM_TOKEN or not CHAT_ID: return
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": CHAT_ID, "text": msg, "parse_mode": "Markdown"}, timeout=15)
    except Exception as e: print(f"❌ Telegram: {e}")

def tg_doc(data, nombre, caption=""):
    if not TELEGRAM_TOKEN or not CHAT_ID: return False
    try:
        r = requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": CHAT_ID, "caption": caption},
            files={"document": (nombre, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30)
        return r.status_code == 200
    except Exception as e: print(f"❌ Telegram doc: {e}"); return False

def _nt(txt): return f"[{NOMBRE_TIENDA}] {txt}"

# ══════════════════════════════════════════════════════════════════════════════
# TEXTO DE AYUDA
# ══════════════════════════════════════════════════════════════════════════════
AYUDA = r"""📋 *Comandos disponibles:*

/menu — Abre el panel de control con botones

*Sincronización*
/sync\_total — Sincroniza precios y stock de todos los productos
/ciclo — Dispara un ciclo de monitoreo manualmente
/registrar\_webhooks — Registra webhooks en Tienda Nube para notificaciones instantáneas
/ver\_webhooks — Muestra los webhooks activos
/borrar\_webhook ID — Elimina un webhook por ID

*Productos*
/listar — Lista los primeros 50 productos del catálogo
/precio Nombre 9999 — Cambia el precio de un producto
/stock Nombre 10 — Cambia el stock de un producto
/ocultar Nombre — Oculta un producto de la tienda
/publicar Nombre — Publica un producto oculto

*Envío gratis*
/fix\_envio\_gratis — Aplica envío gratis a todos los productos
/generar\_descripcion Nombre — Genera descripción con IA (Groq/Llama) para un producto
/generar\_todas\_descripciones — Genera descripciones para todos los productos sin descripción
/set\_video Nombre URL — Agrega video de YouTube a un producto
/generar\_todas\_descripciones — Genera descripciones para todos los productos sin descripción ≥ $100.000 (excepto mesas)
/productos\_sin\_cargar — Productos del proveedor que no tenés en tu tienda (filtrado)
/productos\_sin\_cargar todo — Ídem pero sin filtro (excluye solo Pantallas/Baterías)

*Ofertas*
/ver\_ofertas\_proveedor — Ver todas las ofertas activas del proveedor (numeradas)
/aplicar\_ofertas todos — Aplica todas las ofertas pendientes del proveedor
/aplicar\_ofertas 1 3 — Aplica las ofertas numeradas seleccionadas

*Debug*
/debug\_match Nombre — Verifica el matching de un producto con el proveedor
/debug\_producto ID — Muestra datos crudos de un producto por ID
/debug\_env — Muestra el estado de las variables de entorno

*API / Token*
/estado\_api — Muestra si el token está activo
/borrar\_token — Elimina el token guardado

*Pedidos*
/confirmar\_pedido NUMERO — Confirma y envía al proveedor un pedido con problemas
"""

# ══════════════════════════════════════════════════════════════════════════════
# OAUTH
# ══════════════════════════════════════════════════════════════════════════════
def canjear_code(code):
    try:
        r = requests.post(f"{API_BASE}/oauth/app/token",
            json={"client_id": CLIENT_ID, "client_secret": CLIENT_SECRET,
                  "grant_type": "authorization_code", "code": code},
            headers={"Content-Type":"application/json","User-Agent":USER_AGENT}, timeout=30)
        print(f"OAuth {r.status_code}: {r.text[:200]}")
        if r.status_code in (200, 201):
            d = r.json().get("data", r.json())
            t = d.get("access_token")
            u = str(d.get("store_id") or d.get("user_id") or "")
            if t and u: guardar_token(t, u); return t
    except Exception as e: print(f"❌ OAuth: {e}")
    return None

# ══════════════════════════════════════════════════════════════════════════════
# API TIENDA NEGOCIO
# ══════════════════════════════════════════════════════════════════════════════
def _h(): return {"Authorization":f"Bearer {_token}","User-Agent":USER_AGENT,"Content-Type":"application/json"}

def _get(url, params=None):
    for intento in range(4):
        try:
            r = requests.get(url, headers=_h(), params=params, timeout=40)
            if r.status_code == 429:
                espera = 5 * (intento + 1)
                print(f"   ⚠️ Rate limit GET. Esperando {espera}s..."); time.sleep(espera); continue
            return r
        except requests.exceptions.Timeout:
            print(f"   ⏱️ Timeout GET (intento {intento+1}/4): {url[:60]}")
            time.sleep(3 * (intento + 1))
        except Exception as e:
            print(f"❌ GET: {e}"); time.sleep(2)
    return None

def _put(url, data):
    for intento in range(4):
        try:
            r = requests.put(url, headers=_h(), json=data, timeout=40)
            if r.status_code == 429:
                espera = 15 * (intento + 1)
                print(f"   ⚠️ Rate limit PUT. Esperando {espera}s..."); time.sleep(espera); continue
            time.sleep(0.8)
            return r
        except requests.exceptions.Timeout:
            print(f"   ⏱️ Timeout PUT (intento {intento+1}/4): {url[:60]}")
            time.sleep(5 * (intento + 1))
        except Exception as e:
            print(f"❌ PUT: {e}"); time.sleep(2)
    return None

# ── Caché del catálogo ────────────────────────────────────────────────────────
_cat_cache = None
_cat_ts    = 0

def _fetch_variantes(pid):
    """Descarga variantes de un producto desde la API. Devuelve lista o []."""
    r_var = _get(f"{API_BASE}/products/{pid}/variants", params={"per_page":200})
    variantes = []
    if r_var and r_var.status_code == 200:
        dv = r_var.json()
        lista = dv.get("results", dv) if isinstance(dv,dict) else dv
        if isinstance(lista, list):
            for v in lista:
                vid = v.get("id")
                if not vid: continue
                vals = v.get("values", [])
                vnom = " ".join(str(x.get("es") or x.get("en","")).strip() for x in vals).strip()
                variantes.append({"id":vid,"nombre":vnom,"precio":float(v.get("price",0) or 0)})
    return variantes

def obtener_catalogo(forzar=False, pids_refrescar=None):
    """
    Descarga el catalogo con cache inteligente de variantes en DB.
    - Primera vez / producto nuevo: descarga variantes y las guarda en DB.
    - Ciclos siguientes: usa las variantes cacheadas.
    - pids_refrescar: set de IDs cuyas variantes hay que re-pedir.
    - forzar=True: refresca lista de productos pero usa cache de variantes.
    """
    global _cat_cache, _cat_ts
    if not forzar and _cat_cache and (time.time()-_cat_ts) < 300:
        return _cat_cache
    if not _token: return []

    if pids_refrescar is None:
        pids_refrescar = set()

    db = leer_db()
    var_cache = db.get("variantes_cache", {})  # {str(pid): [lista de variantes]}

    print("📥 Descargando catálogo Tienda Negocio...")
    raw = []; pagina = 1
    while True:
        r = _get(f"{API_BASE}/products", params={"per_page":200,"page":pagina})
        if not r or r.status_code != 200:
            print(f"❌ Catálogo HTTP {r.status_code if r is not None else 'None'}"); break
        data = r.json()
        lote = data.get("results", data) if isinstance(data, dict) else data
        if not lote: break
        raw.extend(lote)
        print(f"   Pág {pagina}: {len(lote)} items (total: {len(raw)})")
        if not (data.get("pagination",{}).get("next_page") if isinstance(data,dict) else None): break
        pagina += 1; time.sleep(0.3)
    print(f"   {len(raw)} productos encontrados.")

    catalogo = []
    nuevas_en_cache = 0
    refrescadas = 0

    for p in raw:
        pid = p.get("id")
        nombre_raw = p.get("name", {})
        nombre = (nombre_raw.get("es") or nombre_raw.get("en") or
                  next(iter(nombre_raw.values()),"")) if isinstance(nombre_raw,dict) else str(nombre_raw or "")
        nombre = nombre.strip()
        if not nombre or not pid: continue

        pid_str = str(pid)
        es_nuevo       = pid_str not in var_cache
        debe_refrescar = pid in pids_refrescar

        if es_nuevo or debe_refrescar:
            variantes = _fetch_variantes(pid)
            var_cache[pid_str] = variantes
            time.sleep(0.4)
            if es_nuevo: nuevas_en_cache += 1
            else:        refrescadas += 1
        else:
            variantes = var_cache[pid_str]

        precio_base = variantes[0]["precio"] if variantes else float(p.get("price",0) or 0)
        catalogo.append({
            "id":pid, "nombre":nombre, "nombre_norm":normalizar(nombre),
            "precio_base":precio_base, "tiene_variantes":len(variantes)>0,
            "variantes":variantes, "published":p.get("published",True),
            "descripcion": ((lambda d: (d.get("es") or d.get("en") or next(iter(d.values()),"")) if isinstance(d, dict) else str(d or ""))(p.get("description",""))).strip()
        })

    # Limpiar del cache productos que ya no existen
    pids_actuales = {str(p.get("id")) for p in raw if p.get("id")}
    eliminados = [k for k in var_cache if k not in pids_actuales]
    for k in eliminados: del var_cache[k]

    db["variantes_cache"] = var_cache
    escribir_db(db)

    info = f"   ✅ Catálogo: {len(catalogo)} productos"
    if nuevas_en_cache: info += f" | {nuevas_en_cache} nuevos cacheados"
    if refrescadas:     info += f" | {refrescadas} variantes refrescadas"
    if eliminados:      info += f" | {len(eliminados)} eliminados del cache"
    print(info)

    _cat_cache = catalogo; _cat_ts = time.time()
    return catalogo

# ── Operaciones API ───────────────────────────────────────────────────────────
def set_precio_variante(vid, precio):
    r = _put(f"{API_BASE}/variants/{vid}", {"price": str(int(precio))})
    ok = r and r.status_code in (200,201)
    if not ok: print(f"  ⚠️ precio variante {vid}: HTTP {r.status_code if r is not None else 'None'}")
    return ok

def set_precio_producto(pid, precio):
    r = _put(f"{API_BASE}/products/{pid}", {"price": str(int(precio))})
    ok = r and r.status_code in (200,201)
    if not ok: print(f"  ⚠️ precio producto {pid}: HTTP {r.status_code if r is not None else 'None'}")
    return ok

def set_stock_variante(vid, stock):
    r = _put(f"{API_BASE}/variants/{vid}", {"stock":int(stock),"stock_management":True})
    return r and r.status_code in (200,201)

def set_stock_producto(pid, stock):
    r = _put(f"{API_BASE}/products/{pid}", {"stock":int(stock),"stock_management":True})
    return r and r.status_code in (200,201)

def set_nombre_producto(pid, nombre):
    r = _put(f"{API_BASE}/products/{pid}", {"name":{"es":nombre}})
    return r and r.status_code in (200,201)

def set_visibilidad(pid, published):
    r = _put(f"{API_BASE}/products/{pid}", {"published":published})
    return r and r.status_code in (200,201)

def set_envio_gratis(pid, activo):
    r = _put(f"{API_BASE}/products/{pid}", {"freeshipping": activo})
    ok = r and r.status_code in (200,201)
    if not ok: print("  Envio gratis HTTP " + str(r.status_code if r is not None else "None"))
    return ok

# ══════════════════════════════════════════════════════════════════════════════
# ENVÍO GRATIS — LÓGICA CENTRALIZADA
# ══════════════════════════════════════════════════════════════════════════════
def _debe_tener_envio_gratis(pid, precio):
    """
    Devuelve True si el producto debe tener envío gratis.
    Regla: precio >= ENVIO_GRATIS_MIN y NO está en la lista de pesados.
    """
    if pid in PRODUCTOS_PESADOS_IDS:
        return False
    return precio >= ENVIO_GRATIS_MIN

def _actualizar_envio_gratis_prod(prod, precio):
    """Activa/desactiva envío gratis según precio. Excluye productos pesados.
    Usa el precio máximo entre variantes para no desactivar por una variante barata."""
    pid = prod["id"]
    if prod.get("variantes"):
        precio_max = max(v["precio"] for v in prod["variantes"])
    else:
        precio_max = precio
    activo = _debe_tener_envio_gratis(pid, precio_max)
    set_envio_gratis(pid, activo)


def run_fix_envio_gratis():
    """
    Comando /fix_envio_gratis — recorre TODO el catálogo y aplica
    la regla de envío gratis independientemente de si el precio cambió.
    Útil para aplicar la lógica por primera vez o después de cambiar ENVIO_GRATIS_MIN.
    """
    if not _token:
        tg("❌ Necesito el token primero."); return

    tg(f"🚚 *{_nt('Fix envío gratis iniciado')}*\nRecorriendo catálogo...")
    catalogo = obtener_catalogo(forzar=True)
    if not catalogo:
        tg("❌ No pude obtener el catálogo."); return

    db   = leer_db()
    prov = db.get("productos_proveedor", {})
    sinc = db.get("sincronizados", {})
    if not prov:
        tg("\u26a0\ufe0f Sin datos del proveedor. Esperá un ciclo primero."); return
    idx = construir_indice(prov)

    activados    = []
    desactivados = []
    pesados_skip = []

    for prod in catalogo:
        pid    = prod["id"]
        nombre = prod["nombre"]

        if pid in PRODUCTOS_PESADOS_IDS:
            pesados_skip.append(f"• *{nombre}* (pesado, sin tocar)")
            continue

        # Usar precio calculado desde el proveedor (no el del catálogo que puede ser 0)
        _, datos_prov = buscar_en_indice(prod["nombre_norm"], idx)
        if datos_prov:
            if datos_prov.get("variantes"):
                precio_max = max(precio_obj(vd["precio"]) for vd in datos_prov["variantes"].values())
            else:
                precio_max = precio_obj(datos_prov["precio_base"])
        else:
            # Sin datos del proveedor: usar precio guardado en sincronizados
            precio_max = sinc.get(nombre, {}).get("precio", 0)

        if precio_max >= ENVIO_GRATIS_MIN:
            if set_envio_gratis(pid, True):
                activados.append(f"• *{nombre}* — ${int(precio_max):,}")
        else:
            if set_envio_gratis(pid, False):
                desactivados.append(f"• *{nombre}* — ${int(precio_max):,}")
        time.sleep(0.5)

    resumen = (f"✅ *{_nt('Fix envío gratis terminado')}*\n\n"
               f"🚚 Activados: *{len(activados)}*\n"
               f"❌ Desactivados: *{len(desactivados)}*\n"
               f"⏭️ Pesados (sin tocar): *{len(pesados_skip)}*")
    tg(resumen)

    if activados:
        for i in range(0, len(activados), 30):
            tg("🚚 *Con envío gratis:*\n\n" + "\n".join(activados[i:i+30]))
    if pesados_skip:
        tg("⏭️ *Productos pesados (cartel manual en foto):*\n\n" + "\n".join(pesados_skip))

# ══════════════════════════════════════════════════════════════════════════════
# SCRAPING PROVEEDOR — API WOOCOMMERCE PÚBLICA
# ══════════════════════════════════════════════════════════════════════════════
SCRAPERAPI_KEY = _e("SCRAPERAPI_KEY")
SCRAPERAPI_URL = "http://api.scraperapi.com"

def _via_scraperapi(url, params=None):
    """Pasa la request por ScraperAPI para evitar bloqueos tipo Cloudflare al proveedor."""
    if not SCRAPERAPI_KEY:
        print("⚠️ Sin SCRAPERAPI_KEY configurada — no puedo usar fallback")
        return None
    target = url
    if params:
        target += "?" + "&".join(f"{k}={v}" for k, v in params.items())
    try:
        r = requests.get(SCRAPERAPI_URL, params={
            "api_key": SCRAPERAPI_KEY,
            "url": target,
        }, timeout=60)
        return r
    except Exception as e:
        print(f"⚠️ ScraperAPI: {e}")
        return None

def _prov_get(url, params=None, usar_scraperapi=False):
    if usar_scraperapi:
        r = _via_scraperapi(url, params)
        return r
    for _ in range(3):
        try:
            if CURL_CFFI_OK:
                # curl-cffi imita TLS fingerprint de Chrome — bypasea Cloudflare
                r = cf_requests.get(url, params=params, timeout=20,
                                    impersonate="chrome120",
                                    headers={"User-Agent": USER_AGENT})
            else:
                r = requests.get(url, params=params, timeout=20,
                                 headers={"User-Agent": USER_AGENT})
            if r.status_code == 429: print("⚠️ Rate limit proveedor"); time.sleep(3); continue
            return r
        except Exception as e: print(f"⚠️ Proveedor API: {e}"); time.sleep(1)
    return None

def _precio_real(p):
    return int(p.get("prices",{}).get("price", 0)) // 100

def _stock_real(p):
    return p.get("add_to_cart",{}).get("maximum") or 0

def scrapear_proveedor():
    productos = {}; pagina = 1; reintentos_202 = 0; via_scraperapi = False
    print("📥 API proveedor...")
    while True:
        r = _prov_get(PROV_API, params={"per_page":100,"page":pagina}, usar_scraperapi=via_scraperapi)
        if not r or r.status_code not in (200, 201, 202):
            print(f"❌ Proveedor HTTP {r.status_code if r is not None else 'None'}"); break
        if r.status_code == 202:
            reintentos_202 += 1
            # ScraperAPI deshabilitado — créditos agotados, se renuevan automáticamente
            # Cuando se renueven, descomentar el bloque de abajo
            # if reintentos_202 == 2 and not via_scraperapi and SCRAPERAPI_KEY:
            #     via_scraperapi = True; continue
            if reintentos_202 >= 5:
                print(f"❌ Proveedor HTTP 202 x{reintentos_202} - abortando scrape")
                tg(f"⚠️ *{NOMBRE_TIENDA}* — Proveedor no responde (HTTP 202 x5{' incluso via ScraperAPI' if via_scraperapi else ''}). Reintentará en el próximo ciclo.")
                break
            espera = 30 if reintentos_202 <= 2 else 60
            print(f"⚠️ Proveedor HTTP 202 ({reintentos_202}/5) - reintentando en {espera}s...")
            time.sleep(espera)
            continue
        if not r.text or not r.text.strip():
            print(f"⚠️ Proveedor respuesta vacía en pág {pagina}, reintentando...")
            time.sleep(5)
            continue
        try:
            lote = r.json()
        except Exception as e:
            print(f"⚠️ Proveedor JSON inválido pág {pagina}: {e}")
            break
        if not lote: break

        for p in lote:
            nombre_orig = p.get("name","").strip()
            if not nombre_orig or len(nombre_orig) < 4: continue
            nombre_base = normalizar(nombre_orig)
            tipo        = p.get("type","simple")
            woo_id      = p.get("id")
            precio_pub      = _precio_real(p)   # precio activo (rebajado si hay oferta)
            precio_original = int(p.get("prices",{}).get("regular_price",0)) // 100
            en_oferta       = p.get("on_sale", False) and 0 < precio_pub < precio_original
            in_stock        = p.get("is_in_stock", False)
            stock_base      = _stock_real(p)

            if precio_pub == 0: continue

            if tipo == "variable":
                variaciones = p.get("variations", [])
                for var_info in variaciones:
                    vid = var_info.get("id")
                    if not vid: continue
                    rv = _prov_get(f"{PROV_API}/{vid}", usar_scraperapi=via_scraperapi)
                    if not rv or rv.status_code != 200: continue
                    vd = rv.json()
                    v_var_str   = vd.get("variation","")
                    v_nombre    = v_var_str.split(":",1)[1].strip() if ":" in v_var_str else v_var_str
                    v_precio    = _precio_real(vd)
                    v_stock     = _stock_real(vd)
                    v_original  = int(vd.get("prices",{}).get("regular_price",0)) // 100
                    v_oferta    = vd.get("on_sale", False) and 0 < v_precio < v_original
                    v_instock   = vd.get("is_in_stock", False)
                    if v_precio == 0: continue

                    clave = normalizar(f"{nombre_orig} ({v_nombre})")
                    productos[clave] = {
                        "nombre_real":           f"{nombre_orig} ({v_nombre})",
                        "nombre_base_proveedor": nombre_base,
                        "precio":                v_precio,
                        "precio_anterior":       v_original if v_oferta else 0,
                        "en_oferta":             v_oferta,
                        "stock":                 v_stock if v_instock else 0,
                        "woo_id":                vid,
                    }
                    time.sleep(0.25)
            else:
                productos[nombre_base] = {
                    "nombre_real":           nombre_orig,
                    "nombre_base_proveedor": nombre_base,
                    "precio":                precio_pub,
                    "precio_anterior":       precio_original if en_oferta else 0,
                    "en_oferta":             en_oferta,
                    "stock":                 stock_base if in_stock else 0,
                    "woo_id":                woo_id,
                }

        print(f"   Pág {pagina}: {len(lote)} prods (total entradas: {len(productos)})")
        if len(lote) < 100: break
        pagina += 1; time.sleep(0.5)

    print(f"   ✅ {len(productos)} entradas del proveedor")
    return productos

# ══════════════════════════════════════════════════════════════════════════════
# MATCHING DE NOMBRES
# ══════════════════════════════════════════════════════════════════════════════
STOP = {'de','para','con','el','la','los','las','un','una','y','en','del','al','a','o','e'}

def normalizar(s):
    return " ".join(re.sub(r'[^a-z0-9 ]',' ', str(s).lower()).split())

def palabras(n):
    return set(n.split()) - STOP

def match(n1, n2):
    if n1 == n2: return True
    w1 = palabras(n1); w2 = palabras(n2)
    if not w1 or not w2: return False
    criticas = {'bateria','battery','bat','face','maneral','mango','zocalo',
                'board','mini','plus','max','kit','ultra','xl',
                'lente','microscopio','trinocular','fuente','estacion','silla','cabina',
                'cam','4k','camara','wifi','pro','lite','set'}
    for c in criticas:
        if (c in w1) != (c in w2): return False
    n1s = {w for w in w1 if any(c.isdigit() for c in w)}
    n2s = {w for w in w2 if any(c.isdigit() for c in w)}
    if n1s and n2s:
        corto = n1s if len(w1) <= len(w2) else n2s
        largo = n2s if len(w1) <= len(w2) else n1s
        if not corto.issubset(largo): return False
    return len(w1 & w2) / min(len(w1), len(w2)) >= 0.75

# ══════════════════════════════════════════════════════════════════════════════
# PRECIO OBJETIVO
# ══════════════════════════════════════════════════════════════════════════════
def precio_obj(costo):
    raw = costo / MARGEN
    if raw >= 100000: return round(raw/1000)*1000
    if raw >= 10000:  return round(raw/500)*500
    if raw >= 1000:   return round(raw/100)*100
    return round(raw/50)*50

# ══════════════════════════════════════════════════════════════════════════════
# ÍNDICE DEL PROVEEDOR
# ══════════════════════════════════════════════════════════════════════════════
def construir_indice(prov):
    idx = {}
    for clave, d in prov.items():
        base = normalizar(d.get("nombre_base_proveedor", clave))
        if base not in idx:
            idx[base] = {"precio_base":d["precio"], "precio_anterior":d.get("precio_anterior",0),
                         "stock_base":d.get("stock",0),
                         "en_oferta":d.get("en_oferta",False), "variantes":{}}

        nombre_real = d.get("nombre_real", clave)
        if '(' in nombre_real:
            var_part = nombre_real.split('(',1)[-1].rstrip(')')
            var_norm = normalizar(var_part)
            if var_norm:
                idx[base]["variantes"][var_norm] = {
                    "precio":          d["precio"],
                    "stock":           d.get("stock",0),
                    "en_oferta":       d.get("en_oferta",False),
                    "precio_anterior": d.get("precio_anterior",0),
                }
        else:
            if d["precio"] < idx[base]["precio_base"]:
                idx[base]["precio_base"]    = d["precio"]
                idx[base]["en_oferta"]      = d.get("en_oferta",False)
                idx[base]["precio_anterior"]= d.get("precio_anterior",0)
            if d.get("stock",0) > idx[base]["stock_base"]:
                idx[base]["stock_base"] = d.get("stock",0)
    return idx

def buscar_en_indice(nombre_norm, idx):
    for base, datos in idx.items():
        if match(nombre_norm, base): return base, datos
    return None, None

# ══════════════════════════════════════════════════════════════════════════════
# SINCRONIZACIÓN DE PRECIOS
# ══════════════════════════════════════════════════════════════════════════════
def sincronizar_precios(prod, datos_prov):
    """
    Actualiza precios de todas las variantes (o el producto si no tiene).
    Si el proveedor tiene oferta activa, usa promotional_price para mostrar
    el precio original tachado junto al precio de oferta.
    """
    pid            = prod["id"]
    variantes      = prod["variantes"]
    en_oferta      = datos_prov.get("en_oferta", False)
    precio_anterior = datos_prov.get("precio_anterior", 0)

    if not variantes:
        p = precio_obj(datos_prov["precio_base"])
        precio_actual = prod.get("precio_base", 0)
        if precio_actual > 0 and p > 0 and (p / precio_actual) < 0.40:
            print("BLOQUEADO baja >60%: " + prod.get("nombre","?")[:40])
            return False, 0, 0

        if en_oferta and precio_anterior > 0:
            # Precio original tachado → precio oferta activo
            p_orig = precio_obj(precio_anterior)
            r = _put(f"{API_BASE}/products/{pid}", {
                "price": str(int(p_orig)),
                "promotional_price": str(int(p))
            })
        else:
            # Sin oferta: precio normal, borrar precio promocional si había
            r = _put(f"{API_BASE}/products/{pid}", {
                "price": str(int(p)),
                "promotional_price": ""
            })
        ok = r and r.status_code in (200, 201)
        if not ok: print(f"  ⚠️ precio producto {pid}: HTTP {r.status_code if r is not None else 'None'}")
        time.sleep(0.4)
        if ok: _actualizar_envio_gratis_prod(prod, p)
        return ok, p, p

    vars_prov = datos_prov.get("variantes", {})
    precios   = {}
    for v in variantes:
        costo = None; costo_ant = 0; v_oferta = False
        vnom = normalizar(v["nombre"])
        for pv_norm, pv_datos in vars_prov.items():
            if match(vnom, pv_norm):
                costo     = pv_datos["precio"]
                costo_ant = pv_datos.get("precio_anterior", 0)
                v_oferta  = pv_datos.get("en_oferta", False)
                break
        if costo is None:
            costo     = datos_prov["precio_base"]
            costo_ant = precio_anterior
            v_oferta  = en_oferta
        precios[v["id"]] = {
            "precio":          precio_obj(costo),
            "precio_anterior": precio_obj(costo_ant) if costo_ant else 0,
            "en_oferta":       v_oferta
        }

    exitos = 0
    for vid, pd in precios.items():
        p = pd["precio"]
        if pd["en_oferta"] and pd["precio_anterior"] > 0:
            data = {"price": str(int(pd["precio_anterior"])), "promotional_price": str(int(p))}
        else:
            data = {"price": str(int(p)), "promotional_price": ""}
        r = _put(f"{API_BASE}/variants/{vid}", data)
        if r and r.status_code in (200, 201): exitos += 1
        else: print(f"  ⚠️ precio variante {vid}: HTTP {r.status_code if r is not None else 'None'}")
        time.sleep(0.4)

    if not precios or exitos == 0: return False, 0, 0

    precio_min = min(pd["precio"] for pd in precios.values())
    precio_max = max(pd["precio"] for pd in precios.values())
    _actualizar_envio_gratis_prod(prod, precio_min)
    return True, precio_min, precio_max

# ══════════════════════════════════════════════════════════════════════════════
# SINCRONIZACIÓN DE STOCK REAL
# ══════════════════════════════════════════════════════════════════════════════
def sincronizar_stock(prod, datos_prov, sinc):
    if not _token: return
    nombre_real = prod["nombre"]
    pid         = prod["id"]
    variantes   = prod["variantes"]
    vars_prov   = datos_prov.get("variantes", {})
    stock_base  = datos_prov.get("stock_base", 0)

    if stock_base >= 9999: return

    if variantes:
        for v in variantes:
            vnom  = normalizar(v["nombre"])
            stock = stock_base
            for pv_norm, pv_datos in vars_prov.items():
                if match(vnom, pv_norm):
                    stock = pv_datos.get("stock", stock_base); break
            ultimo = sinc.get(nombre_real,{}).get(f"stock_{v['id']}")
            if ultimo != stock:
                set_stock_variante(v["id"], stock)
                sinc.setdefault(nombre_real,{})[f"stock_{v['id']}"] = stock
            time.sleep(0.3)
    else:
        ultimo = sinc.get(nombre_real,{}).get("stock_sinc")
        if ultimo != stock_base:
            set_stock_producto(pid, stock_base)
            sinc.setdefault(nombre_real,{})["stock_sinc"] = stock_base

# ══════════════════════════════════════════════════════════════════════════════
# MARCAR SIN STOCK
# ══════════════════════════════════════════════════════════════════════════════
def marcar_sin_stock(prod):
    pid = prod["id"]; variantes = prod["variantes"]
    if variantes:
        ok = True
        for v in variantes:
            if not set_stock_variante(v["id"], 0): ok = False
            time.sleep(0.3)
        return ok
    else:
        r = _put(f"{API_BASE}/products/{pid}", {"stock":0,"stock_management":True})
        if not (r and r.status_code in (200,201)):
            print(f"⚠️ No pude poner stock=0 en producto {pid} — queda visible")
            return True
        return True

# ══════════════════════════════════════════════════════════════════════════════
# SYNC TOTAL
# ══════════════════════════════════════════════════════════════════════════════
def run_sync_total():
    db   = leer_db()
    prov = db.get("productos_proveedor",{})
    sinc = db.get("sincronizados",{})
    if not prov:
        tg("⚠️ Sin datos del proveedor. Esperá un ciclo de monitoreo primero."); return

    catalogo = obtener_catalogo(forzar=True)
    if not catalogo:
        tg("❌ No pude obtener el catálogo de la API."); return
    db = leer_db()  # Re-leer para no pisar variantes_cache
    sinc = db.get("sincronizados", sinc)

    idx   = construir_indice(prov)
    total = len(catalogo)
    act_precios   = []
    act_sin_stock = []
    errores       = []

    tg(f"🔄 *{_nt('Sync total iniciada')}*\n{total} productos a procesar...")

    for i, prod in enumerate(catalogo):
        nombre_real = prod["nombre"]
        precio_web  = prod["precio_base"]
        _, datos_prov = buscar_en_indice(prod["nombre_norm"], idx)

        if datos_prov is None:
            if marcar_sin_stock(prod):
                sinc[nombre_real] = {"sin_stock": True}
                act_sin_stock.append(f"• *{nombre_real}*")
            else:
                errores.append(nombre_real)
        else:
            if sinc.get(nombre_real,{}).get("sin_stock"):
                precio_guardado = sinc[nombre_real].get("precio",0)
                sinc[nombre_real] = {"precio":precio_guardado} if precio_guardado else {}

            ok, p_min, p_max = sincronizar_precios(prod, datos_prov)
            if ok:
                sinc.setdefault(nombre_real,{})["precio"] = p_min
                if int(precio_web) != p_min:
                    rango = f"${p_min:,}" if p_min==p_max else f"${p_min:,}–${p_max:,}"
                    etiq  = f"({len(prod['variantes'])} var.)" if prod["tiene_variantes"] else "(sin var.)"
                    act_precios.append(f"• *{nombre_real}* {etiq}\n  Antes: ${int(precio_web):,} → *Nuevo: {rango}*")
            else:
                errores.append(nombre_real)

            sincronizar_stock(prod, datos_prov, sinc)

        if (i+1) % 50 == 0: print(f"   Sync [{i+1}/{total}]")

    db["sincronizados"] = sinc
    escribir_db(db)

    resumen = (f"✅ *{_nt('Sync total terminada')}*\n\n"
               f"📊 Total: *{total}*\n"
               f"💲 Precios actualizados: *{len(act_precios)}*\n"
               f"📦 Sin stock: *{len(act_sin_stock)}*\n")
    if errores: resumen += f"⚠️ Errores: *{len(errores)}*"
    tg(resumen)

    for i in range(0, len(act_precios), 20):
        tg(f"💲 *Precios actualizados:*\n\n" + "\n\n".join(act_precios[i:i+20]))
    for i in range(0, len(act_sin_stock), 30):
        tg(f"📦 *Marcados sin stock:*\n\n" + "\n".join(act_sin_stock[i:i+30]))

# ══════════════════════════════════════════════════════════════════════════════
# GMAIL
# ══════════════════════════════════════════════════════════════════════════════
def chequear_gmail():
    if not GMAIL_USER or not GMAIL_PASS: return []
    pedidos = []; conn = None
    try:
        conn = imaplib.IMAP4_SSL("imap.gmail.com")
        conn.login(GMAIL_USER, GMAIL_PASS); conn.select("inbox")
        st, msgs = conn.search(None, '(UNSEEN FROM "tiendanegocio.com")')
        if st != "OK" or not msgs[0]: conn.close(); conn.logout(); return []
        for mid in msgs[0].split():
            sid = mid.decode()
            res, data = conn.fetch(mid, "(RFC822)")
            if res != "OK": continue
            msg = email.message_from_bytes(data[0][1])
            subj, enc = decode_header(msg["Subject"])[0]
            if isinstance(subj, bytes): subj = subj.decode(enc or "utf-8")
            if not any(p in subj.lower() for p in ["compra","realiz","pedido","venta"]): continue
            num_m = re.search(r"#(\d+)", subj)
            num = num_m.group(1) if num_m else sid
            cuerpo = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type()=="text/plain":
                        cuerpo = part.get_payload(decode=True).decode("utf-8","ignore"); break
            else:
                cuerpo = msg.get_payload(decode=True).decode("utf-8","ignore")
            items = _parsear_items(cuerpo)
            if items:
                pedidos.append({"id":sid,"num":num,"items":items})
                conn.store(mid, "+FLAGS", "\\Seen")
        conn.close(); conn.logout()
    except Exception as e:
        print("Gmail: " + str(e))
        try:
            if conn: conn.close(); conn.logout()
        except Exception: pass
    return pedidos

def _parsear_items(cuerpo):
    items=[]; en=False
    for l in cuerpo.split('\n'):
        l=l.strip()
        if "productos:" in l.lower(): en=True; continue
        if en:
            if not l or "subtotal:" in l.lower(): en=False; continue
            if l.startswith('-'):
                m=re.match(r'-\s*(.+?)\s+x(\d+)\s*-',l)
                if m: items.append({"nombre":m.group(1),"cant":int(m.group(2))})
    return items

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def generar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "❌ Falta openpyxl en requirements.txt"
    db  = leer_db(); prov = db.get("productos_proveedor",{})
    if not prov: return None, "❌ Sin datos del proveedor."
    catalogo = obtener_catalogo()
    if not catalogo: return None, "❌ Sin catálogo de la API."
    idx = construir_indice(prov)
    cols = ["Hash","Nombre del producto","Precio","Oferta","Stock",
            "Visibilidad (Visible o Oculto)","Descripción","SKU",
            "Peso en KG","Alto en CM","Ancho en CM","Profundidad en CM",
            "Nombre de variante #1","Opción de variante #1",
            "Nombre de variante #2","Opción de variante #2",
            "Nombre de variante #3","Opción de variante #3",
            "Categorías > Subcategorías > … > Subcategorías"]
    filas=[]; act=ign=sin=0
    for prod in catalogo:
        _,dp = buscar_en_indice(prod["nombre_norm"], idx)
        if dp is None: sin+=1; continue
        p = precio_obj(dp["precio_base"])
        if p == int(prod["precio_base"]): ign+=1; continue
        f={c:"" for c in cols}
        f["Hash"]=prod["nombre_norm"]; f["Nombre del producto"]=prod["nombre"]; f["Precio"]=p
        filas.append(f); act+=1
    if not filas: return None, f"ℹ️ Sin cambios.\n• {ign} ya correctos\n• {sin} sin match"
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Productos"
    ws.append(cols)
    hf=PatternFill("solid",fgColor="1F6B3B"); hfnt=Font(bold=True,color="FFFFFF",name="Arial",size=10)
    for c in ws[1]: c.fill=hf; c.font=hfnt; c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=25
    fills=[PatternFill("solid",fgColor="F0F7F2"),PatternFill("solid",fgColor="FFFFFF")]
    fn=Font(name="Arial",size=9); fv=Font(name="Arial",size=9,bold=True,color="1F6B3B")
    for i,fila in enumerate(filas,2):
        ws.append([fila[c] for c in cols])
        for cell in ws[i]: cell.fill=fills[i%2]; cell.font=fn
        ws.cell(row=i,column=3).font=fv; ws.cell(row=i,column=3).number_format='#,##0'
    for col,w in zip("ABCDEFGHIJKLMNOPQRS",[38,48,12,8,8,12,8,8,8,8,8,8,20,22,20,22,20,22,30]):
        ws.column_dimensions[col].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue(),(f"✅ Excel:\n• *{act}* a actualizar\n• *{ign}* ya correctos\n• *{sin}* sin match\n\n"
                           f"Importalo: *Productos → Importar y exportar → Importar*")

# ══════════════════════════════════════════════════════════════════════════════
# CICLO DE MONITOREO
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# DETECCIÓN DE ÓRDENES VIA API (más confiable que Gmail)
# ══════════════════════════════════════════════════════════════════════════════
def chequear_ordenes_api():
    """
    Consulta GET /orders buscando pedidos nuevos desde el último conocido.
    No depende de emails leídos. Se ejecuta en cada ciclo.
    Primera ejecución: guarda el estado actual sin notificar (evita spam de órdenes viejas).
    """
    if not _token: return []
    try:
        db  = leer_db()
        ultimo_id = db.get("ultimo_orden_id", 0)

        params = {"per_page": 50}
        if ultimo_id:
            params["since_id"] = ultimo_id

        r = _get(f"{API_BASE}/orders", params=params)
        if r is None or r.status_code != 200:
            detalle = r.text[:150] if r is not None else "Sin respuesta"
            print(f"⚠️ Órdenes API: HTTP {r.status_code if r is not None else 'None'} — {detalle}")
            return []

        data    = r.json()
        ordenes = data.get("results", data) if isinstance(data, dict) else data
        if not isinstance(ordenes, list) or not ordenes:
            return []

        max_id = max(o.get("id", 0) for o in ordenes)

        # Primera ejecución o DB reseteada: guardar estado sin notificar.
        # Marcamos TODAS las ordenes existentes como procesadas
        # para evitar spam tras un redeploy de Railway.
        if ultimo_id == 0:
            db["ultimo_orden_id"] = max_id
            ya_proc = db.get("pedidos_procesados", [])
            for o in ordenes:
                oid = str(o.get("id", ""))
                if oid and oid not in ya_proc:
                    ya_proc.append(oid)
            db["pedidos_procesados"] = ya_proc
            escribir_db(db)
            print(f"   Ordenes API: primer ciclo — {len(ordenes)} ordenes marcadas como procesadas")
            return []

        # Actualizar ultimo ID conocido
        if max_id > ultimo_id:
            db["ultimo_orden_id"] = max_id
            escribir_db(db)

        print(f"   Ordenes API: {len(ordenes)} desde ID {ultimo_id}")
        return ordenes

    except Exception as e:
        print(f"❌ chequear_ordenes_api: {e}")
        return []


def notificar_orden_nueva(orden):
    """Arma y envía notificación Telegram para una orden nueva."""
    num      = str(orden.get("number", orden.get("id", "?")))
    productos = orden.get("products", [])
    total    = float(orden.get("total") or 0)
    pago_st  = orden.get("payment_status", "?")
    envio    = orden.get("shipping", {})
    metodo   = envio.get("pickup_type", "") or envio.get("method", "") or ""

    db   = leer_db()
    prov = db.get("productos_proveedor", {})

    msg  = "🛒 *" + _nt("¡Nuevo Pedido #" + num + "!") + "*\n\n"
    msg += f"💰 Total: *${total:,.0f}*\n"
    msg += f"💳 Pago: *{pago_st}*\n"
    if metodo:
        msg += f"📦 Envío: *{metodo}*\n"
    msg += "\n"

    if prov:
        idx = construir_indice(prov)
        for item in productos:
            nombre   = item.get("name", "?")
            cantidad = item.get("quantity", 1)
            _, dp    = buscar_en_indice(normalizar(nombre), idx)
            if dp:
                stock = dp.get("stock_base", 0)
                icono = "⚠️" if stock < cantidad else "✅"
                extra = f" (stock proveedor: {stock})" if stock < cantidad else ""
                msg += f"{icono} *{nombre}* x{cantidad}{extra}\n"
            else:
                msg += f"❓ *{nombre}* x{cantidad}\n"
    else:
        for item in productos:
            msg += f"• *{item.get('name','?')}* x{item.get('quantity',1)}\n"

    tg(msg)

def ciclo_monitoreo():
    print(f"\n─── 🔄 Ciclo {datetime.now().strftime('%H:%M')} ───")
    db          = leer_db()
    prov_ant    = db.get("productos_proveedor",{})
    sinc        = db.get("sincronizados",{})
    ped_proc    = db.get("pedidos_procesados",[])

    # Detección de órdenes via API (principal) + Gmail (backup)
    for orden in chequear_ordenes_api():
        oid = str(orden.get("id", ""))
        if oid not in ped_proc:
            notificar_orden_nueva(orden)
            ped_proc.append(oid)

    # Guardar ped_proc YA para no perderlo si el ciclo aborta por fallo del proveedor
    db["pedidos_procesados"] = ped_proc
    escribir_db(db)

    # Gmail como backup (por si acaso)
    for ped in chequear_gmail():
        if ped["id"] not in ped_proc:
            ped_proc.append(ped["id"])

    prov_nuevo = scrapear_proveedor()
    if not prov_nuevo:
        print("⚠️ Proveedor 0 productos. Abortando."); return

    prov_consolidado = {**prov_ant, **prov_nuevo}
    idx      = construir_indice(prov_nuevo)

    # Leer pids marcados para refrescar variantes en este ciclo
    pids_refrescar = set(db.get("pids_refrescar", []))
    db["pids_refrescar"] = []  # Limpiar para el proximo ciclo

    catalogo = obtener_catalogo(pids_refrescar=pids_refrescar) if _token else []
    db = leer_db()  # Re-leer para no pisar variantes_cache guardado por obtener_catalogo
    sinc = db.get("sincronizados", sinc)

    bloque_nuevos = ""; bloque_recuperados = ""
    lineas_precios    = []
    lineas_sin_stock  = []
    lineas_stock_bajo = []

    # ── Analizar novedades del proveedor ──────────────────────────────────────
    ofertas_nuevas = {}
    for clave, datos in prov_nuevo.items():
        viejo = prov_ant.get(clave, {})

        # Oferta nueva real del proveedor
        if datos["en_oferta"] and datos.get("precio_anterior",0) > datos["precio"] and not viejo.get("en_oferta", False):
            base = datos["nombre_base_proveedor"]
            if base not in ofertas_nuevas:
                ofertas_nuevas[base] = datos

        # Producto nuevo que no tengo en mi web
        base_norm = datos.get("nombre_base_proveedor", clave)
        tengo = any(match(p["nombre_norm"], base_norm) for p in catalogo) if catalogo else False
        if not tengo and not viejo:
            p_sug = precio_obj(datos["precio"])
            if any(w in clave for w in PALABRAS_INTERES):
                bloque_nuevos += (f"• *{datos['nombre_real']}*\n"
                                  f"  Costo: ${datos['precio']:,} → Sugerido: ${p_sug:,}\n\n")

        # Stock recuperado
        if viejo and not viejo.get("stock", True) and datos.get("stock", True):
            nombre_r = datos["nombre_real"]
            if sinc.get(nombre_r,{}).get("sin_stock"):
                precio_guardado = sinc[nombre_r].get("precio",0)
                sinc[nombre_r]  = {"precio":precio_guardado} if precio_guardado else {}
                bloque_recuperados += f"• *{nombre_r}*\n\n"

        # Alerta stock bajo
        stock = datos.get("stock", 0)
        if isinstance(stock, (int,float)) and 0 < stock <= ALERTA_STOCK and stock < 9999:
            nombre_r = datos["nombre_real"]
            ultimo_alerta_stock = sinc.get(nombre_r,{}).get("alerta_stock_cant")
            if ultimo_alerta_stock != int(stock):
                lineas_stock_bajo.append(f"• *{nombre_r}*: {int(stock)} unidad{'es' if stock>1 else ''}")
                sinc.setdefault(nombre_r,{})["alerta_stock_cant"] = int(stock)

    # Guardar y notificar ofertas pendientes
    if ofertas_nuevas:
        db["ofertas_pendientes"] = ofertas_nuevas
        nums = [f"{i+1}. *{d['nombre_real']}*\n   Reg: ${d.get('precio_anterior',0):,} → 🔥 ${d['precio']:,}"
                for i,(k,d) in enumerate(ofertas_nuevas.items())]
        tg(f"🏷️ *{_nt('Ofertas nuevas del Proveedor')}*\n\n" + "\n".join(nums) +
           f"\n\nUsá `/aplicar_ofertas todos` o `/aplicar_ofertas 1 3` para aplicarlas.")

    # ── Actualizar mi tienda ──────────────────────────────────────────────────
    for prod in catalogo:
        nombre_real = prod["nombre"]
        precio_web  = prod["precio_base"]
        sync_actual = sinc.get(nombre_real, {})
        _, datos_prov = buscar_en_indice(prod["nombre_norm"], idx)

        if datos_prov is None:
            if not sync_actual.get("sin_stock", False):
                if marcar_sin_stock(prod):
                    sinc[nombre_real] = {**sync_actual, "sin_stock": True}
                    etiq = f"({len(prod['variantes'])} var.)" if prod["tiene_variantes"] else "(sin var.)"
                    lineas_sin_stock.append(f"• *{nombre_real}* {etiq}")
        else:
            if sync_actual.get("sin_stock"):
                precio_guardado = sync_actual.get("precio", 0)
                sinc[nombre_real] = {"precio": precio_guardado} if precio_guardado else {}
                sync_actual = sinc[nombre_real]

            p_obj_calc = precio_obj(datos_prov["precio_base"])
            if p_obj_calc != sync_actual.get("precio", 0):
                ratio = (p_obj_calc / precio_web) if precio_web > 0 else 1
                if ratio < 0.40:
                    alerta = "[" + NOMBRE_TIENDA + "] PRECIO BLOQUEADO baja >60%: "
                    alerta += nombre_real + " actual $" + str(int(precio_web))
                    alerta += " calculado $" + str(p_obj_calc)
                    alerta += " | verificar /debug_match"
                    tg(alerta)
                else:
                    ok, p_min, p_max = sincronizar_precios(prod, datos_prov)
                    if ok:
                        sinc.setdefault(nombre_real, {})["precio"] = p_min
                        # Marcar para refrescar variantes en el proximo ciclo
                        db.setdefault("pids_refrescar", [])
                        if prod["tiene_variantes"] and prod["id"] not in db["pids_refrescar"]:
                            db["pids_refrescar"].append(prod["id"])
                        if int(precio_web) != p_min:
                            rango = str(p_min) if p_min == p_max else str(p_min) + "-" + str(p_max)
                            nvar = len(prod["variantes"])
                            etiq = "(" + str(nvar) + " var.)" if prod["tiene_variantes"] else "(sin var.)"
                            linea = nombre_real + " " + etiq
                            linea += " | Antes $" + str(int(precio_web))
                            linea += " Nuevo $" + rango
                            lineas_precios.append(linea)
            else:
                # Precio no cambió → igual verificar envío gratis por si no estaba aplicado
                _actualizar_envio_gratis_prod(prod, p_obj_calc)

            sincronizar_stock(prod, datos_prov, sinc)

    # ── Notificaciones ────────────────────────────────────────────────────────
    if bloque_nuevos:
        tg(f"🔥 *{_nt('¡Nuevo producto en el Proveedor!')}*\n\n{bloque_nuevos}")
    if bloque_recuperados:
        tg(f"🔄 *{_nt('¡Stock recuperado en Proveedor!')}*\n\n{bloque_recuperados}")
    for i in range(0, len(lineas_precios), 20):
        tg(f"💲 *{_nt('Precios actualizados:')}*\n\n" + "\n\n".join(lineas_precios[i:i+20]))
    for i in range(0, len(lineas_sin_stock), 30):
        tg(f"📦 *{_nt('Marcados sin stock:')}*\n\n" + "\n".join(lineas_sin_stock[i:i+30]))
    if lineas_stock_bajo:
        tg(f"⚠️ *{_nt('Stock bajo en proveedor:')}*\n\n" + "\n".join(lineas_stock_bajo))

    if not any([bloque_nuevos, bloque_recuperados, lineas_precios, lineas_sin_stock, lineas_stock_bajo]):
        print("✅ Sin cambios.")

    db["productos_proveedor"] = prov_consolidado
    db["sincronizados"]       = sinc
    db["pedidos_procesados"]  = ped_proc
    db["api_token"]           = _token
    db["api_user_id"]         = _store_id
    escribir_db(db)
    print("─── ✅ Ciclo completado ───")

# ══════════════════════════════════════════════════════════════════════════════
# COMPRAS AUTOMÁTICAS AL PROVEEDOR
# ══════════════════════════════════════════════════════════════════════════════
def _prov_session():
    if not PROV_USER or not PROV_PASS:
        print("❌ Falta PROV_USER o PROV_PASS en Railway Variables")
        return None, None
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    try:
        r = session.get(PROV_LOGIN_URL, timeout=20)
        if r.status_code != 200:
            print(f"❌ Login page HTTP {r.status_code}"); return None, None
        login_data = {
            "log":         PROV_USER,
            "pwd":         PROV_PASS,
            "wp-submit":   "Acceder",
            "redirect_to": "https://rxzweb.com/mi-cuenta/",
            "testcookie":  "1"
        }
        r2 = session.post(PROV_LOGIN_URL, data=login_data, timeout=20)
        if "mi-cuenta" not in r2.url and "dashboard" not in r2.url:
            print(f"❌ Login falló. URL final: {r2.url[:80]}")
            return None, None
        print("✅ Login rxzweb exitoso")
        r3 = session.get(PROV_CART_URL, timeout=20)
        nonce = r3.headers.get("X-WC-Store-API-Nonce") or r3.headers.get("Nonce", "")
        if not nonce:
            r4 = session.get("https://rxzweb.com/", timeout=20)
            import re as _re
            m = _re.search(r'"nonce":"([^"]+)"', r4.text)
            if m: nonce = m.group(1)
        print(f"   Nonce: {nonce[:20] if nonce else 'NO ENCONTRADO'}...")
        return session, nonce
    except Exception as e:
        print(f"❌ Session: {e}"); return None, None

def _prov_add_to_cart(session, nonce, product_id, variation_id, quantity):
    headers = {"Nonce": nonce, "Content-Type": "application/json"}
    payload = {"id": variation_id or product_id, "quantity": quantity}
    try:
        r = session.post(f"{PROV_CART_URL}/add-item", json=payload, headers=headers, timeout=20)
        ok = r.status_code in (200, 201)
        print(f"  {'✅' if ok else '❌'} Carrito add {product_id}: HTTP {r.status_code}")
        if not ok: print(f"     {r.text[:100]}")
        return ok
    except Exception as e:
        print(f"❌ Add to cart: {e}"); return False

def _prov_checkout(session, nonce, datos_cliente):
    headers = {"Nonce": nonce, "Content-Type": "application/json"}
    payload = {
        "billing_address": {
            "first_name": datos_cliente.get("first_name", ""),
            "last_name":  datos_cliente.get("last_name", ""),
            "company":    CUIT_PROVEEDOR or "",
            "address_1":  datos_cliente.get("address_1", ""),
            "city":       datos_cliente.get("city", ""),
            "state":      datos_cliente.get("state", "B"),
            "postcode":   datos_cliente.get("postcode", ""),
            "country":    "AR",
            "email":      datos_cliente.get("email", PROV_USER or ""),
            "phone":      datos_cliente.get("phone", ""),
        },
        "shipping_address": {
            "first_name": datos_cliente.get("first_name", ""),
            "last_name":  datos_cliente.get("last_name", ""),
            "company":    "",
            "address_1":  datos_cliente.get("address_1", ""),
            "city":       datos_cliente.get("city", ""),
            "state":      datos_cliente.get("state", "B"),
            "postcode":   datos_cliente.get("postcode", ""),
            "country":    "AR",
            "phone":      datos_cliente.get("phone", ""),
        },
        "payment_method": "cod",
        "customer_note":  datos_cliente.get("nota", ""),
    }
    try:
        r = session.post(PROV_CHKOUT_URL, json=payload, headers=headers, timeout=30)
        if r.status_code in (200, 201):
            data = r.json()
            orden = data.get("order_id") or data.get("id") or data.get("number")
            print(f"✅ Pedido creado en proveedor: #{orden}")
            return str(orden)
        else:
            print(f"❌ Checkout HTTP {r.status_code}: {r.text[:200]}")
            return None
    except Exception as e:
        print(f"❌ Checkout: {e}"); return None

def hacer_pedido_proveedor(items, datos_cliente, nota=""):
    print("🛒 Iniciando compra automática al proveedor...")
    session, nonce = _prov_session()
    if not session:
        tg("❌ No pude iniciar sesión en el proveedor. Verificá PROV_USER y PROV_PASS.")
        return None
    for item in items:
        ok = _prov_add_to_cart(session, nonce, item["product_id"], item.get("variation_id"), item["quantity"])
        if not ok:
            tg(f"❌ No pude agregar al carrito: {item.get('nombre', item['product_id'])}")
            return None
        time.sleep(0.5)
    if nota:
        datos_cliente["nota"] = nota
    orden = _prov_checkout(session, nonce, datos_cliente)
    return orden

def procesar_orden_pagada(datos_orden):
    num_orden  = str(datos_orden.get("id") or datos_orden.get("number","?"))
    cliente    = datos_orden.get("billing", {})
    productos  = datos_orden.get("products", datos_orden.get("line_items", []))
    envio_tipo = datos_orden.get("shipping", {}).get("pickup_type", "shipping")
    es_retiro  = "pickup" in str(envio_tipo).lower() or "retiro" in str(envio_tipo).lower()

    db      = leer_db()
    prov    = db.get("productos_proveedor", {})
    idx     = construir_indice(prov)

    items_pedido = []
    hay_problema = False
    lineas_msg   = ["Orden pagada #" + num_orden, ""]

    for item in productos:
        nombre   = item.get("name", item.get("product_name","?"))
        cantidad = item.get("quantity", 1)
        nn       = normalizar(nombre)
        _, dp    = buscar_en_indice(nn, idx)

        if dp is None:
            lineas_msg.append("NO ENCONTRADO: " + nombre + " x" + str(cantidad))
            hay_problema = True
            continue

        stock_disp = dp.get("stock_base", 0)
        if stock_disp < cantidad:
            lineas_msg.append("STOCK BAJO: " + nombre + " x" + str(cantidad) + " (disponible: " + str(stock_disp) + ")")
            hay_problema = True
        else:
            lineas_msg.append("OK: " + nombre + " x" + str(cantidad))

        woo_id = None; var_id = None
        for clave, d in prov.items():
            base = normalizar(d.get("nombre_base_proveedor", clave))
            if match(nn, base):
                woo_id = d.get("woo_id")
                if "(" in d.get("nombre_real",""):
                    var_id = woo_id
                    for c2,d2 in prov.items():
                        if normalizar(d2.get("nombre_base_proveedor",""))==base and "(" not in d2.get("nombre_real",""):
                            woo_id=d2.get("woo_id"); break
                break

        items_pedido.append({"nombre":nombre,"product_id":woo_id,"variation_id":var_id,"quantity":cantidad})

    ship = datos_orden.get("shipping", {})
    datos_cli = {
        "first_name": cliente.get("first_name",""),
        "last_name":  cliente.get("last_name",""),
        "address_1":  ship.get("address", cliente.get("address","")),
        "city":       ship.get("city", cliente.get("city","")),
        "state":      ship.get("province","B"),
        "postcode":   ship.get("zipcode",""),
        "phone":      cliente.get("phone",""),
        "email":      cliente.get("email",""),
    }
    nota = "Orden cliente #" + num_orden + (" - RETIRO EN LOCAL" if es_retiro else "")

    sep = chr(10)
    if hay_problema:
        lineas_msg.append("")
        lineas_msg.append("Hay problemas de stock.")
        lineas_msg.append("Usa /confirmar_pedido " + num_orden + " para proceder igual.")
        tg("[" + NOMBRE_TIENDA + "] " + sep.join(lineas_msg))
        db["pedido_pendiente"] = {"num_orden":num_orden,"items":items_pedido,"cliente":datos_cli,"nota":nota}
        escribir_db(db)
    else:
        tg("[" + NOMBRE_TIENDA + "] " + sep.join(lineas_msg))
        orden_prov = hacer_pedido_proveedor(items_pedido, datos_cli, nota)
        if orden_prov:
            tg("[" + NOMBRE_TIENDA + "] Pedido enviado al proveedor. Cliente #" + num_orden + " -> Proveedor #" + orden_prov)
            db.setdefault("ordenes",{})[num_orden] = {"orden_prov": orden_prov}
            escribir_db(db)
        else:
            tg("[" + NOMBRE_TIENDA + "] ERROR: No pude hacer el pedido #" + num_orden + ". Hacelo manualmente.")

# ══════════════════════════════════════════════════════════════════════════════
# COMANDOS TELEGRAM
# ══════════════════════════════════════════════════════════════════════════════
def run_productos_sin_cargar(modo_todo=False):
    """
    Muestra productos del proveedor que no están cargados en la tienda.
    Excluye siempre: Pantallas/Modulos y Baterias.
    modo_todo=False → filtra por PALABRAS_INTERES
    modo_todo=True  → muestra todo (menos categorías excluidas)
    """
    if not _token:
        tg("❌ Necesito el token primero."); return

    tg(f"🔍 *{_nt('Buscando productos del proveedor sin cargar...')}*")

    # Scrapear proveedor fresco con categorías
    productos_prov = {}; pagina = 1
    while True:
        r = _prov_get(PROV_API, params={"per_page":100,"page":pagina})
        if not r or r.status_code not in (200,201,202): break
        lote = r.json()
        if not lote: break
        for p in lote:
            nombre = p.get("name","").strip()
            if not nombre or len(nombre) < 4: continue
            # Chequear categorías excluidas
            cats = [c.get("slug","").lower() + " " + c.get("name","").lower()
                    for c in p.get("categories",[])]
            cats_str = " ".join(cats)
            if any(exc in cats_str for exc in CATEGORIAS_EXCLUIDAS):
                continue
            precio = int(p.get("prices",{}).get("price",0)) // 100
            if precio == 0: continue
            stock = p.get("add_to_cart",{}).get("maximum") or 0
            productos_prov[normalizar(nombre)] = {
                "nombre_real": nombre,
                "precio": precio,
                "precio_sug": precio_obj(precio),
                "stock": stock,
            }
        if len(lote) < 100: break
        pagina += 1; time.sleep(0.5)

    catalogo = obtener_catalogo()
    cat_norms = {p["nombre_norm"] for p in catalogo}

    # Filtrar los que no están en mi tienda
    sin_cargar = []
    for norm, datos in productos_prov.items():
        en_tienda = any(match(norm, cn) for cn in cat_norms)
        if en_tienda: continue
        if not modo_todo:
            if not any(w in norm for w in PALABRAS_INTERES): continue
        sin_cargar.append(datos)

    if not sin_cargar:
        modo_txt = "en toda la lista" if modo_todo else "con palabras de interés"
        tg(f"ℹ️ No hay productos sin cargar {modo_txt}."); return

    modo_txt = "🗂 *Todos* (sin Pantallas/Baterías)" if modo_todo else "🎯 *Filtrado por palabras de interés*"
    header = f"📦 *{_nt('Productos sin cargar en tu tienda')}*\n{modo_txt}\nTotal: *{len(sin_cargar)}*"
    tg(header)

    lineas = []
    for d in sin_cargar:
        stock_txt = f"Stock: {d['stock']}" if d['stock'] < 9999 else "Stock: ∞"
        linea = f"• *{d['nombre_real']}*\n  Costo: ${d['precio']:,} → Sugerido: ${d['precio_sug']:,} | {stock_txt}"
        lineas.append(linea)

    for i in range(0, len(lineas), 20):
        tg("\n\n".join(lineas[i:i+20]))



# ══════════════════════════════════════════════════════════════════════════════
# WEBHOOK SERVER — Recibe eventos de Tienda Nube en tiempo real
# ══════════════════════════════════════════════════════════════════════════════
if FLASK_OK:
    flask_app = Flask(__name__)

    @flask_app.route("/webhook", methods=["POST"])
    def recibir_webhook():
        try:
            data = request.get_json(silent=True) or {}
            evento   = data.get("event", "")
            orden_id = data.get("id")
            print(f"📨 Webhook: {evento} id={orden_id}")
            if orden_id:
                if evento == "order/created":
                    threading.Thread(target=procesar_webhook_pedido_nuevo, args=(orden_id,), daemon=True).start()
                elif evento == "order/paid":
                    threading.Thread(target=procesar_webhook_pedido_pagado, args=(orden_id,), daemon=True).start()
        except Exception as e:
            print(f"❌ Webhook error: {e}")
        return jsonify({"status": "ok"}), 200

    @flask_app.route("/health", methods=["GET"])
    def health():
        return jsonify({"status": "ok", "tienda": NOMBRE_TIENDA}), 200

    def run_flask():
        port = int(os.environ.get("PORT", 8080))
        print(f"🌐 Webhook server en puerto {port}")
        flask_app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)


def obtener_orden(orden_id):
    """Obtiene datos completos de una orden desde la API de Tienda Nube."""
    r = _get(f"{API_BASE}/orders/{orden_id}")
    if r and r.status_code == 200:
        return r.json()
    print(f"❌ No pude obtener orden {orden_id}: HTTP {r.status_code if r is not None else 'None'}")
    return None


def procesar_webhook_pedido_nuevo(orden_id):
    """
    Webhook order/created: notifica al instante con items y chequeo de stock.
    Reemplaza la detección por Gmail para pedidos nuevos.
    """
    time.sleep(1)  # Espera breve para que TN termine de guardar
    orden = obtener_orden(orden_id)
    if not orden:
        tg("⚠️ *" + _nt('Nuevo pedido') + f"* recibido pero no pude obtener detalles (ID: {orden_id})")
        return

    num      = str(orden.get("number", orden_id))
    productos = orden.get("products", [])
    total    = orden.get("total", "0")
    pago_st  = orden.get("payment_status", "?")
    envio    = orden.get("shipping", {})
    metodo   = envio.get("pickup_type", "") or envio.get("method", "")

    db   = leer_db()
    prov = db.get("productos_proveedor", {})

    msg = "🛒 *" + _nt('¡Nuevo Pedido #' + num + '!') + "*\n\n"
    msg += f"💰 Total: *${float(total or 0):,.0f}*\n"
    msg += f"💳 Pago: *{pago_st}*\n"
    if metodo:
        msg += f"📦 Envío: *{metodo}*\n"
    msg += "\n"

    if prov:
        idx = construir_indice(prov)
        for item in productos:
            nombre   = item.get("name", "?")
            cantidad = item.get("quantity", 1)
            nn       = normalizar(nombre)
            _, dp    = buscar_en_indice(nn, idx)
            if dp:
                stock = dp.get("stock_base", 0)
                icono = "⚠️" if stock < cantidad else "✅"
                stock_txt = f" (stock: {stock})" if stock < cantidad else ""
                msg += f"{icono} *{nombre}* x{cantidad}{stock_txt}\n"
            else:
                msg += f"❓ *{nombre}* x{cantidad}\n"
    else:
        for item in productos:
            msg += f"• *{item.get('name','?')}* x{item.get('quantity',1)}\n"

    tg(msg)


def procesar_webhook_pedido_pagado(orden_id):
    """
    Webhook order/paid: notifica pago confirmado y dispara compra automática.
    """
    time.sleep(1)
    orden = obtener_orden(orden_id)
    if not orden:
        tg("⚠️ *" + _nt('Pedido pagado') + f"* pero no pude obtener detalles (ID: {orden_id})")
        return

    num = str(orden.get("number", orden_id))
    tg("💚 *" + _nt('¡Pago confirmado! Pedido #' + num) + "*")
    procesar_orden_pagada(orden)



# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE DESCRIPCIONES CON IA
# ══════════════════════════════════════════════════════════════════════════════
def generar_descripcion_ia(nombre, precio):
    """Llama a Groq API (Llama) para generar descripcion de producto. Gratis."""
    if not GROQ_API_KEY:
        return None, "Falta GROQ_API_KEY en Railway Variables"
    prompt = (
        "Sos un experto en herramientas para tecnicos de reparacion electronica de celulares y placas en Argentina.\n\n"
        f"Genera una descripcion de producto para una tienda online orientada a tecnicos profesionales.\n"
        f"Producto: {nombre}\n"
        f"Precio: ${precio:,}\n\n"
        "La descripcion debe:\n"
        "- Estar en espanol argentino, sin tuteo\n"
        "- Tener entre 100 y 200 palabras\n"
        "- Explicar para que sirve y que problemas resuelve\n"
        "- Mencionar compatibilidades o especificaciones si aplica\n"
        "- Usar lenguaje tecnico apropiado para profesionales\n"
        "- NO usar markdown ni asteriscos, solo texto plano con saltos de linea\n"
        "- NO inventar especificaciones que no conoces\n\n"
        "Solo devuelve la descripcion, sin titulo ni comentarios adicionales."
    )
    try:
        r = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_API_KEY}",
                     "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile",
                  "max_tokens": 1000,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=30
        )
        if r.status_code == 200:
            return r.json()["choices"][0]["message"]["content"].strip(), None
        return None, f"Groq API HTTP {r.status_code}: {r.text[:150]}"
    except Exception as e:
        return None, f"Error Groq API: {e}"


def set_descripcion_producto(pid, descripcion):
    """Sube la descripcion a Tienda Nube."""
    r = _put(f"{API_BASE}/products/{pid}", {"description": descripcion})
    return r is not None and r.status_code in (200, 201)


def run_generar_todas_descripciones():
    """
    Recorre el catalogo y genera descripciones para productos que no tienen.
    Envia progreso cada 10 productos.
    """
    catalogo = obtener_catalogo(forzar=True)
    if not catalogo:
        tg("No pude obtener el catalogo."); return

    sin_desc = [p for p in catalogo if not p.get("descripcion")]
    total = len(sin_desc)

    if not sin_desc:
        tg("Todos los productos ya tienen descripcion."); return

    tg(f"*{_nt('Generando descripciones')}*\n\n"
       f"Productos sin descripcion: *{total}*\n"
       f"Tiempo estimado: ~{total * 3 // 60} minutos\n\n"
       f"Te aviso cada 10 productos.")

    ok = 0; errores = []; i = 0
    for prod in sin_desc:
        i += 1
        desc, err = generar_descripcion_ia(prod["nombre"], int(prod["precio_base"]))
        if err:
            errores.append(prod["nombre"])
            print(f"Error IA {prod['nombre']}: {err}")
        elif set_descripcion_producto(prod["id"], desc):
            ok += 1
            print(f"OK desc: {prod['nombre']}")
        else:
            errores.append(prod["nombre"])
        if i % 10 == 0:
            tg(f"Progreso: *{i}/{total}* ({ok} OK, {len(errores)} errores)")
        time.sleep(2)  # Evitar rate limits

    tg(f"*{_nt('Descripciones completadas')}*\n\n"
       f"Exitosas: *{ok}*\n"
       f"Errores: *{len(errores)}*" +
       (f"\n\nFallaron:\n" + "\n".join(f"• {e}" for e in errores[:10]) if errores else ""))



def run_ver_ofertas_proveedor():
    """
    Muestra TODAS las ofertas activas del proveedor en este momento.
    Las enumera para poder aplicar todas o solo algunas con /aplicar_ofertas.
    """
    db   = leer_db()
    prov = db.get("productos_proveedor", {})
    if not prov:
        tg("Sin datos del proveedor. Esperá un ciclo primero."); return

    # Filtrar solo los que tienen oferta activa
    ofertas = {}
    for clave, d in prov.items():
        if not d.get("en_oferta"): continue
        precio     = d.get("precio", 0)
        precio_ant = d.get("precio_anterior", 0)
        if not precio or not precio_ant or precio >= precio_ant: continue
        # Usar nombre base del proveedor para agrupar (evitar duplicados de variantes)
        base = d.get("nombre_base_proveedor", clave)
        if base not in ofertas:
            ofertas[base] = {
                "nombre_real":    d.get("nombre_real", base),
                "precio":         precio,
                "precio_anterior": precio_ant,
                "precio_web":     precio_obj(precio),
                "precio_web_ant": precio_obj(precio_ant),
                "descuento":      round((1 - precio / precio_ant) * 100),
            }

    if not ofertas:
        tg("No hay ofertas activas del proveedor en este momento."); return

    lista = list(ofertas.items())
    lineas = []
    for i, (base, d) in enumerate(lista, 1):
        ahorro = d["precio_web_ant"] - d["precio_web"]
        lineas.append(
            f"*{i}.* {d['nombre_real']}\n"
            f"   Costo: ~~${d['precio_anterior']:,}~~ → ${d['precio']:,} (-{d['descuento']}%)\n"
            f"   Tu precio: ~~${d['precio_web_ant']:,}~~ → *${d['precio_web']:,}* (ahorrás ${ahorro:,})"
        )

    # Guardar en DB para que /aplicar_ofertas pueda usarlas
    db["ofertas_pendientes"] = {base: {
        "nombre_real":    d["nombre_real"],
        "precio":         d["precio"],
        "precio_anterior": d["precio_anterior"],
        "en_oferta":      True,
        "stock":          prov.get(base, {}).get("stock", 0),
        "nombre_base_proveedor": base,
    } for base, d in lista}
    escribir_db(db)

    encabezado = (f"*{_nt('Ofertas activas del proveedor')}*\n"
                  f"Total: *{len(lista)}*\n\n"
                  f"Usá `/aplicar_ofertas todos` o `/aplicar_ofertas 1 3 5` para aplicar.")

    tg(encabezado)
    for i in range(0, len(lineas), 15):
        tg("\n\n".join(lineas[i:i+15]))


def set_video_producto(pid, url):
    """Sube URL de video YouTube a Tienda Nube."""
    r = _put(f"{API_BASE}/products/{pid}", {"video_url": url})
    return r is not None and r.status_code in (200, 201)


def tg_menu():
    """Envía menú interactivo con botones inline agrupados por categoría."""
    if not TELEGRAM_TOKEN or not CHAT_ID: return
    keyboard = {
        "inline_keyboard": [
            [{"text": "🔄 Sync Total", "callback_data": "/sync_total"},
             {"text": "🔁 Ciclo Manual", "callback_data": "/ciclo"}],
            [{"text": "📋 Listar Productos", "callback_data": "/listar"},
             {"text": "📦 Sin Cargar", "callback_data": "/productos_sin_cargar todo"}],
            [{"text": "🏷️ Ver Ofertas Proveedor", "callback_data": "/ver_ofertas_proveedor"},
             {"text": "✅ Aplicar Ofertas", "callback_data": "/aplicar_ofertas todos"}],
            [{"text": "🚚 Fix Envío Gratis", "callback_data": "/fix_envio_gratis"},
             {"text": "📊 Exportar Precios", "callback_data": "/exportar_precios"}],
            [{"text": "✍️ Generar Descripciones", "callback_data": "/generar_todas_descripciones"}],
            [{"text": "🔗 Ver Webhooks", "callback_data": "/ver_webhooks"},
             {"text": "📡 Registrar Webhooks", "callback_data": "/registrar_webhooks"}],
            [{"text": "🔑 Estado API", "callback_data": "/estado_api"},
             {"text": "🛠️ Debug Env", "callback_data": "/debug_env"}],
            [{"text": "📄 Ayuda completa", "callback_data": "/ayuda"}],
        ]
    }
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": CHAT_ID,
                  "text": f"*{_nt('Panel de control')}*\nElegí una opción:",
                  "parse_mode": "Markdown",
                  "reply_markup": keyboard},
            timeout=15
        )
    except Exception as e:
        print(f"❌ Menu: {e}")


def procesar_cmd(texto):
    global _token, _store_id
    texto = texto.strip()

    if "?code=" in texto:
        code = texto.split("?code=")[1].split("&")[0].strip()
        tg("🔄 Canjeando código OAuth...")
        token = canjear_code(code)
        if token:
            tg(f"✅ *¡Token obtenido!*\n\nGuardá en Railway → Variables:\n"
               f"• `API_TOKEN` = `{token}`\n• `API_USER_ID` = `{_store_id}`")
        else:
            tg("❌ Token fallido. El código dura 1 minuto.")
        return

    cmd = texto.lower().split()
    if not cmd: return

    if cmd[0] == "/menu":
        tg_menu()

    elif cmd[0] == "/ayuda":
        tg(AYUDA)

    elif cmd[0] == "/estado_api":
        if _token:
            tg(f"✅ *Token activo*\nStore ID: `{_store_id}`\nToken: `{_token[:12]}...`")
        else:
            tg("❌ Sin token. Mandá el `?code=` para obtenerlo.")

    elif cmd[0] == "/borrar_token":
        _token = _store_id = None
        db = leer_db(); db.pop("api_token",None); db.pop("api_user_id",None)
        escribir_db(db); tg("🗑️ Token eliminado.")

    elif cmd[0] == "/listar":
        if not _token: tg("❌ Necesito el token primero."); return
        tg("⏳ Cargando catálogo...")
        catalogo = obtener_catalogo(forzar=True)
        if not catalogo: tg("No encontré productos."); return
        lineas = []
        for p in catalogo[:50]:
            cv   = len(p["variantes"])
            icon = "✅" if p["published"] else "🚫"
            extra = f"({cv} var.)" if cv > 0 else "(sin var.)"
            lineas.append(icon + " *" + p["nombre"] + "* " + extra + " — $" + str(int(p["precio_base"])) + " (ID:" + str(p["id"]) + ")")
        msg = f"📦 *{len(catalogo)} productos:*\n\n" + "\n".join(lineas)
        if len(catalogo) > 50: msg += f"\n\n_...y {len(catalogo)-50} más_"
        tg(msg)

    elif cmd[0] == "/sync_total":
        if not _token: tg("❌ Necesito el token primero."); return
        threading.Thread(target=run_sync_total, daemon=True).start()

    elif cmd[0] == "/fix_envio_gratis":
        if not _token: tg("❌ Necesito el token primero."); return
        threading.Thread(target=run_fix_envio_gratis, daemon=True).start()

    elif cmd[0] == "/ocultar":
        if not _token: tg("❌ Necesito el token primero."); return
        nombre = " ".join(texto.split()[1:])
        if not nombre: tg("Uso: `/ocultar Nombre`"); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre),p["nombre_norm"])),None)
        if prod:
            if set_visibilidad(prod["id"],False): tg(f"🚫 *{prod['nombre']}* ocultado.")
            else: tg("❌ No pude ocultar.")
        else: tg(f"❌ No encontré *{nombre}*.")

    elif cmd[0] == "/publicar":
        if not _token: tg("❌ Necesito el token primero."); return
        nombre = " ".join(texto.split()[1:])
        if not nombre: tg("Uso: `/publicar Nombre`"); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre),p["nombre_norm"])),None)
        if prod:
            if set_visibilidad(prod["id"],True): tg(f"✅ *{prod['nombre']}* publicado.")
            else: tg("❌ No pude publicar.")
        else: tg(f"❌ No encontré *{nombre}*.")

    elif cmd[0] == "/stock":
        if not _token: tg("❌ Necesito el token primero."); return
        partes = texto.split()
        if len(partes) < 3: tg("Uso: `/stock Nombre 10`"); return
        try: nuevo=int(partes[-1]); nombre=" ".join(partes[1:-1])
        except ValueError: tg("El último parámetro debe ser un número."); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre),p["nombre_norm"])),None)
        if prod:
            if prod["variantes"]:
                ok = all(set_stock_variante(v["id"],nuevo) for v in prod["variantes"])
            else:
                ok = set_stock_producto(prod["id"],nuevo)
            tg(f"📦 *{prod['nombre']}* → stock *{nuevo}*." if ok else "❌ No pude actualizar.")
        else: tg(f"❌ No encontré *{nombre}*.")

    elif cmd[0] == "/precio":
        if not _token: tg("❌ Necesito el token primero."); return
        partes = texto.split()
        if len(partes) < 3: tg("Uso: `/precio Nombre 9999`"); return
        try: nuevo=int(partes[-1]); nombre=" ".join(partes[1:-1])
        except ValueError: tg("El último parámetro debe ser un número."); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre),p["nombre_norm"])),None)
        if prod:
            if prod["variantes"]:
                ok = all(set_precio_variante(v["id"],nuevo) for v in prod["variantes"])
            else:
                ok = set_precio_producto(prod["id"],nuevo)
            tg(f"💲 *{prod['nombre']}* → *${nuevo:,}*." if ok else "❌ No pude actualizar.")
        else: tg(f"❌ No encontré *{nombre}*.")

    elif cmd[0] == "/exportar_precios":
        tg("⏳ Generando Excel...")
        data, msg = generar_excel()
        if data:
            fecha = datetime.now().strftime("%d-%m-%Y")
            if not tg_doc(data, f"precios_{fecha}.xlsx", caption=msg):
                tg("❌ Excel generado pero falló el envío.")
        else: tg(msg)

    elif cmd[0] == "/ciclo":
        tg(f"🔄 *{_nt('Ciclo manual iniciado')}*")
        threading.Thread(target=ciclo_monitoreo, daemon=True).start()

    elif cmd[0] == "/debug_ordenes":
        if not _token: tg("❌ Necesito el token primero."); return
        tg("🔄 Probando GET /orders directo...")
        try:
            r = requests.get(f"{API_BASE}/orders", headers=_h(), params={"per_page":5}, timeout=20)
            tg(f"✅ Respuesta recibida\nHTTP {r.status_code}\n`{r.text[:300]}`")
        except Exception as e:
            tg(f"❌ Excepción real: `{type(e).__name__}: {e}`")

    elif cmd[0] == "/set_video":
        if not _token: tg("Necesito el token primero."); return
        partes = texto.split()
        if len(partes) < 3: tg("Uso: `/set_video Nombre del producto https://youtube.com/...`"); return
        # Último token es la URL, el resto es el nombre
        url = partes[-1]
        nombre = " ".join(partes[1:-1])
        if "youtube.com" not in url and "youtu.be" not in url:
            tg("La URL debe ser de YouTube."); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre), p["nombre_norm"])), None)
        if not prod:
            tg(f"No encontre *{nombre}* en tu catalogo."); return
        if set_video_producto(prod["id"], url):
            tg(f"Video agregado a *{prod['nombre']}*")
        else:
            tg("No pude agregar el video.")

    elif cmd[0] == "/generar_descripcion":
        if not _token: tg("Necesito el token primero."); return
        if not GROQ_API_KEY:
            tg("Falta `GROQ_API_KEY` en Railway Variables."); return
        nombre = " ".join(texto.split()[1:]).strip()
        if not nombre: tg("Uso: `/generar_descripcion Nombre del producto`"); return
        prod = next((p for p in obtener_catalogo() if match(normalizar(nombre), p["nombre_norm"])), None)
        if not prod:
            tg(f"No encontre *{nombre}* en tu catalogo."); return
        tg(f"Generando descripcion para *{prod['nombre']}*...")
        desc, err = generar_descripcion_ia(prod["nombre"], int(prod["precio_base"]))
        if err:
            tg(f"Error: {err}"); return
        if set_descripcion_producto(prod["id"], desc):
            tg(f"*{prod['nombre']}*\n\n{desc}")
        else:
            tg("No pude subir la descripcion a Tienda Nube.")

    elif cmd[0] == "/generar_todas_descripciones":
        if not _token: tg("Necesito el token primero."); return
        if not GROQ_API_KEY:
            tg("Falta `GROQ_API_KEY` en Railway Variables."); return
        threading.Thread(target=run_generar_todas_descripciones, daemon=True).start()
        tg(f"Iniciando generacion masiva de descripciones...")

    elif cmd[0] == "/test_scraperapi":
        if not SCRAPERAPI_KEY:
            tg("❌ Falta `SCRAPERAPI_KEY` en Railway Variables."); return
        tg("🔄 Probando ScraperAPI contra el proveedor...")
        r = _via_scraperapi(PROV_API, params={"per_page":1,"page":1})
        if r and r.status_code == 200:
            try:
                data = r.json()
                nombre = data[0].get("name","?") if data else "?"
                tg(f"✅ ScraperAPI funciona. Producto de prueba: *{nombre}*")
            except Exception as e:
                tg(f"⚠️ HTTP 200 pero no pude parsear JSON: {e}")
        else:
            tg(f"❌ ScraperAPI HTTP {r.status_code if r is not None else 'Sin respuesta'}")

    elif cmd[0] == "/registrar_webhooks":
        if not _token: tg("❌ Necesito el token primero."); return
        if not WEBHOOK_BASE_URL:
            tg("❌ Falta `WEBHOOK_BASE_URL` en Railway Variables.\n"
               "Valor: `https://bot-dropshipping-production.up.railway.app`"); return
        tg("🔄 Registrando webhooks en Tienda Nube...")
        resultados = []
        for evento in ["order/created", "order/paid"]:
            try:
                r = requests.post(f"{API_BASE}/webhooks", headers=_h(),
                    json={"event": evento, "url": f"{WEBHOOK_BASE_URL}/webhook"}, timeout=30)
                if r is not None and r.status_code in (200, 201):
                    resultados.append(f"✅ `{evento}`")
                else:
                    codigo = r.status_code if r is not None else "Sin respuesta"
                    detalle = r.text[:100] if r is not None else ""
                    resultados.append(f"❌ `{evento}` — HTTP {codigo}: {detalle}")
            except Exception as e:
                resultados.append(f"❌ `{evento}` — Error: {e}")
        tg("*Webhooks:*\n" + "\n".join(resultados))

    elif cmd[0] == "/ver_webhooks":
        if not _token: tg("❌ Necesito el token primero."); return
        r = _get(f"{API_BASE}/webhooks")
        if r and r.status_code == 200:
            hooks = r.json()
            results = hooks.get("results", hooks) if isinstance(hooks, dict) else hooks
            if not results:
                tg("ℹ️ No hay webhooks registrados."); return
            lineas = [f"• `{h.get('event')}` (ID:{h.get('id')})\n  → {h.get('url')}" for h in results]
            tg("🔗 *Webhooks activos:*\n\n" + "\n\n".join(lineas))
        else:
            tg(f"❌ HTTP {r.status_code if r is not None else 'None'}")

    elif cmd[0] == "/borrar_webhook":
        if not _token: tg("❌ Necesito el token primero."); return
        wid = texto.split()[1] if len(texto.split()) > 1 else ""
        if not wid: tg("Uso: /borrar_webhook ID"); return
        r = requests.delete(f"{API_BASE}/webhooks/{wid}", headers=_h(), timeout=30)
        tg(f"✅ Webhook {wid} borrado." if r and r.status_code in (200,204) else f"❌ HTTP {r.status_code if r is not None else 'None'}")

    elif cmd[0] == "/productos_sin_cargar":
        if not _token: tg("❌ Necesito el token primero."); return
        modo_todo = len(cmd) > 1 and cmd[1] == "todo"
        threading.Thread(target=run_productos_sin_cargar, args=(modo_todo,), daemon=True).start()

    elif cmd[0] == "/ver_ofertas_proveedor":
        threading.Thread(target=run_ver_ofertas_proveedor, daemon=True).start()

    elif cmd[0] == "/aplicar_ofertas":
        db = leer_db()
        ofertas = db.get("ofertas_pendientes", {})
        if not ofertas:
            tg("ℹ️ No hay ofertas pendientes de aplicar."); return
        lista = list(ofertas.items())
        partes = texto.split()[1:]
        if not partes:
            tg("Uso: `/aplicar_ofertas todos` o `/aplicar_ofertas 1 3`"); return
        if partes[0] == "todos":
            seleccion = list(range(len(lista)))
        else:
            seleccion = []
            for x in partes:
                try:
                    n = int(x) - 1
                    if 0 <= n < len(lista): seleccion.append(n)
                except ValueError: pass
        if not seleccion:
            tg("❌ Número inválido."); return
        catalogo = obtener_catalogo()
        idx = construir_indice({k:v for k,v in ofertas.items()})
        aplicadas = []
        for i in seleccion:
            base_norm, datos_prov = lista[i]
            prod = next((p for p in catalogo if match(p["nombre_norm"], base_norm)), None)
            if not prod:
                aplicadas.append(f"• *{datos_prov['nombre_real']}* — No encontrado en tu tienda"); continue
            ok, p_min, p_max = sincronizar_precios(prod, datos_prov)
            if ok:
                rango = f"${p_min:,}" if p_min==p_max else f"${p_min:,}–${p_max:,}"
                aplicadas.append(f"• *{prod['nombre']}*\n  Precio oferta: {rango}")
            else:
                aplicadas.append(f"• *{prod['nombre']}* — Error al actualizar")
        db["ofertas_pendientes"] = {}
        escribir_db(db)
        tg("✅ *Ofertas aplicadas:*\n\n" + "\n\n".join(aplicadas))

    elif cmd[0] == "/debug_match":
        nombre = " ".join(texto.split()[1:]).strip()
        if not nombre: tg("Uso: /debug_match Nombre del producto"); return
        nombre_norm = normalizar(nombre)
        db = leer_db(); prov = db.get("productos_proveedor",{})
        if not prov: tg("Sin datos del proveedor."); return
        idx = construir_indice(prov)
        base_norm, datos_prov = buscar_en_indice(nombre_norm, idx)
        lineas = ["Debug: " + nombre, ""]
        if datos_prov:
            p_o = precio_obj(datos_prov["precio_base"])
            lineas += ["PROVEEDOR: encontrado",
                       "  Base: " + base_norm,
                       "  Precio base: $" + str(datos_prov["precio_base"]) + " -> Web: $" + str(p_o),
                       "  Stock base: " + str(datos_prov.get("stock_base","?")),
                       "  Variantes prov: " + str(len(datos_prov["variantes"]))]
            for vn, vd in list(datos_prov["variantes"].items())[:5]:
                lineas.append("    [" + vn[:35] + "]: $" + str(vd["precio"]) + " st:" + str(vd.get("stock","?")))
        else:
            lineas.append("PROVEEDOR: NO encontrado")
        lineas.append("")
        catalogo = obtener_catalogo() if _token else []
        prod = next((p for p in catalogo if match(nombre_norm,p["nombre_norm"])),None)
        if prod and datos_prov:
            vars_prov_idx = datos_prov.get("variantes", {})
            lineas += ["CATALOGO API: encontrado",
                       "  ID: " + str(prod["id"]),
                       "  Nombre: " + prod["nombre"],
                       "  Variantes API: " + str(len(prod["variantes"])),
                       "  Precio actual: $" + str(int(prod["precio_base"])),
                       "  Matching variantes:"]
            for v in prod["variantes"][:8]:
                vnom = normalizar(v["nombre"])
                match_prov = next(((pv,pd) for pv,pd in vars_prov_idx.items() if match(vnom,pv)), None)
                if match_prov:
                    p_calc = precio_obj(match_prov[1]["precio"])
                    lineas.append("    OK " + v["nombre"][:30] + " -> $" + str(p_calc))
                else:
                    p_fb = precio_obj(datos_prov["precio_base"])
                    lineas.append("    FB " + v["nombre"][:30] + " -> $" + str(p_fb))
        elif prod:
            lineas += ["CATALOGO API: encontrado (sin datos prov)",
                       "  ID: " + str(prod["id"]),
                       "  Nombre: " + prod["nombre"],
                       "  Variantes: " + str(len(prod["variantes"])),
                       "  Precio actual: $" + str(int(prod["precio_base"]))]
        else:
            lineas.append("CATALOGO API: " + ("NO encontrado" if _token else "sin token"))
        tg("\n".join(lineas))

    elif cmd[0] == "/debug_producto":
        if not _token: tg("Sin token."); return
        pid = texto.split()[1] if len(texto.split()) > 1 else ""
        if not pid: tg("Uso: /debug_producto ID"); return
        r = _get(f"{API_BASE}/products/{pid}")
        if not r or r.status_code != 200: tg("HTTP " + str(r.status_code if r is not None else 0)); return
        d = r.json()
        sep = chr(10)
        info = [k + ": " + str(v) for k,v in d.items() if not isinstance(v,(dict,list))]
        tg("Producto " + str(pid) + ":" + sep + sep.join(info[:25]))

    elif cmd[0] == "/debug_env":
        t = os.environ.get("API_TOKEN",""); u = os.environ.get("API_USER_ID","")
        lineas = [
            "TELEGRAM_TOKEN: " + ("OK" if TELEGRAM_TOKEN else "NO"),
            "CHAT_ID: "        + ("OK" if CHAT_ID else "NO"),
            "CLIENT_ID: "      + ("OK" if CLIENT_ID else "NO"),
            "API_TOKEN env: "  + (t[:10]+"..." if t else "NO"),
            "API_USER_ID env: "+ (u if u else "NO"),
            "Token memoria: "  + (str(_token)[:10]+"..." if _token else "NONE"),
            "UserID memoria: " + (_store_id or "NONE"),
            "NOMBRE_TIENDA: "  + NOMBRE_TIENDA,
            "MARGEN: "         + str(MARGEN),
        ]
        tg("Diagnóstico:\n" + "\n".join(lineas))

    elif cmd[0] == "/confirmar_pedido":
        partes = texto.split()
        if len(partes) < 2:
            tg("Uso: /confirmar_pedido NUMERO_ORDEN"); return
        num = partes[1]
        db  = leer_db()
        pp  = db.get("pedido_pendiente", {})
        if not pp or str(pp.get("num_orden")) != str(num):
            tg(f"No hay pedido pendiente #{num}."); return
        tg(f"🔄 Confirmando pedido #{num} al proveedor...")
        orden_prov = hacer_pedido_proveedor(pp["items"], pp["cliente"], pp.get("nota",""))
        if orden_prov:
            tg(f"✅ Pedido #{num} enviado al proveedor como #{orden_prov}")
            db.pop("pedido_pendiente", None)
            db.setdefault("ordenes", {})[str(num)] = {"orden_prov": orden_prov}
            escribir_db(db)
        else:
            tg(f"❌ No pude hacer el pedido #{num}. Intentá manualmente.")

    else:
        tg("❓ Comando no reconocido. Mandá `/ayuda`.")

# ── Loop Telegram ─────────────────────────────────────────────────────────────
def escuchar_telegram():
    if not TELEGRAM_TOKEN: return
    offset = 0; url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
    print("📡 Telegram activo...")
    while True:
        try:
            r = requests.get(f"{url}?offset={offset}&timeout=10", timeout=15)
            if r.status_code == 200:
                for u in r.json().get("result",[]):
                    offset = u["update_id"] + 1
                    # Mensaje de texto normal
                    msg = u.get("message",{}); texto = msg.get("text","")
                    if str(msg.get("chat",{}).get("id","")) == CHAT_ID and texto:
                        print(f"📨 {texto[:60]}")
                        try: procesar_cmd(texto)
                        except Exception as e: print(f"❌ Cmd: {e}")
                    # Callback de botón inline
                    cb = u.get("callback_query",{})
                    if cb and str(cb.get("message",{}).get("chat",{}).get("id","")) == CHAT_ID:
                        cb_data = cb.get("data","")
                        cb_id   = cb.get("id","")
                        # Responder al callback para quitar el "reloj" del botón
                        try:
                            requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/answerCallbackQuery",
                                json={"callback_query_id": cb_id}, timeout=5)
                        except Exception: pass
                        if cb_data:
                            print(f"🔘 Botón: {cb_data[:60]}")
                            try: procesar_cmd(cb_data)
                            except Exception as e: print(f"❌ Botón cmd: {e}")
        except Exception as e: print(f"⚠️ Telegram: {e}")
        time.sleep(1)

# ══════════════════════════════════════════════════════════════════════════════
# ARRANQUE
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("🚀 Bot iniciado...")
    print(f"   API_TOKEN env:   {'SÍ (' + os.environ.get('API_TOKEN','')[:8] + '...)' if os.environ.get('API_TOKEN') else 'NO'}")
    print(f"   API_USER_ID env: {os.environ.get('API_USER_ID','NO')}")
    print(f"   NOMBRE_TIENDA:   {NOMBRE_TIENDA}")
    print(f"   MARGEN:          {MARGEN} ({round((1-MARGEN)*100)}% ganancia)")

    cargar_token()

    if _token:
        tg(f"🟢 *{_nt('Bot iniciado')}* — Token activo (store_id: `{_store_id}`)\n\n"
           f"Mandá `/menu` para abrir el panel de control con botones.\n"
           f"Mandá `/ayuda` para ver todos los comandos.")
    else:
        tg(f"🟡 *{_nt('Bot iniciado')}* — Sin token API.\n"
           f"Mandá `/debug_env` para diagnosticar.")

    threading.Thread(target=escuchar_telegram, daemon=True).start()
    if FLASK_OK:
        threading.Thread(target=run_flask, daemon=True).start()
    else:
        print('⚠️ Flask no disponible — webhooks desactivados')
    try:
        ciclo_monitoreo()
    except Exception as e:
        print(f"⚠️ Error en ciclo inicial: {e}")
        tg(f"⚠️ *{_nt('Error en ciclo inicial')}*\n`{e}`\nEl bot sigue activo, reintentará en {CICLO_MINUTOS} min.")
    while True:
        time.sleep(CICLO_MINUTOS * 60)
        try:
            ciclo_monitoreo()
        except Exception as e:
            print(f"⚠️ Error en ciclo: {e}")
            tg(f"⚠️ *{_nt('Error en ciclo')}*\n`{e}`\nReintentando en {CICLO_MINUTOS} min.")
